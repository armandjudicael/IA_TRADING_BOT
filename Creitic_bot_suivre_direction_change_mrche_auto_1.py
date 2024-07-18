import os
import pandas as pd
import ta
from iqoptionapi.stable_api import IQ_Option
import time
import datetime
import threading
from colorama import init, Fore, Style
from openpyxl import Workbook
from openpyxl.styles import Alignment

class TradingBot:
    def __init__(self):
        self.email = "judicael.ratombotiana@gmail.com"
        self.password = "Aj!30071999@jv"
        self.account_type = input("Type de compte (demo/real): ").strip().lower()
        self.available_pairs = {}
        self.solde_initial = 0.00
        self.excel_filename = 'trading_results.xlsx'

        try:
            self.api = IQ_Option(self.email, self.password)
            connect_result, connect_message = self.api.connect()
            print(Fore.RED + f"Connexion : {connect_result}" + Style.RESET_ALL)
            if connect_result:
                print(Fore.YELLOW + "Connexion établie, récupération des paires disponibles..." + Style.RESET_ALL)
                self.available_pairs = self.fetch_available_pairs_with_payouts()
                self.solde_initial = self.api.get_balance()
                self.api.api.close()  # Close the connection after fetching the pairs
                print(Fore.GREEN + "Connexion fermée après récupération des paires." + Style.RESET_ALL)
            else:
                print(Fore.RED + f"Erreur de connexion : {connect_message}" + Style.RESET_ALL)
        except Exception as e:
            print(Fore.RED + f"Exception lors de la connexion : {e}" + Style.RESET_ALL)

        print("\nPaires disponibles :")
        for idx, pair in enumerate(self.available_pairs):
            print(f"{idx + 1}. {pair}")

        pair_index = int(input("\nSélectionnez le numéro de la paire : ")) - 1
        self.pair = self.available_pairs[pair_index].strip().upper()

        self.stake = float(input("Mise initiale: "))
        self.martingale = float(input("Martingale: "))
        self.current_stake = self.stake
        self.action_ = "null"
        self.action = "default"
        self.status = ""
        self.total_profit = 0.00
        self.profit = 0.00
        self.profit_aff = 0.00
        self.id = 0
        self.recuperation_martingale = 0
        self.somme_marge_perdu = 0
        self.result = "default"
        self.time = "default"
        self.initial = "true"
        self.initial_testing_action = "default"

        # Initialize Excel workbook and worksheet
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Trading Results'

        # Write headers to Excel sheet
        headers = ['Timestamp', 'Pair', 'Action', 'Stake', 'Martingale', 'Result', 'Profit', 'Total Profit']
        for col_num, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=1, column=col_num)
            cell.value = header
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Save the workbook initially
        self.wb.save(self.excel_filename)

    def fetch_available_pairs_with_payouts(self):
        self.api.connect()
        pairs = self.api.get_all_open_time()
        available_pairs = [pair for pair in pairs['digital'] if pairs['digital'][pair]['open']]
        return available_pairs

    def get_payout(self, pair):
        instruments = self.api.get_all_init()
        if "instruments" in instruments and "digital-option" in instruments["instruments"]:
            digital_options = instruments["instruments"]["digital-option"]
            for option in digital_options:
                if option["active_id"] == self.api.get_name_by_activeId(pair)["active_id"]:
                    return option["profit"]["commission"]
        return None

    def is_pair_open(self, pair):
        self.api.connect()
        pairs = self.api.get_all_open_time()
        return pairs['digital'][pair]['open']

    def select_new_pair(self):
        open_pairs = [pair for pair in self.available_pairs if self.is_pair_open(pair)]
        if open_pairs:
            self.pair = open_pairs[0]
            print(f"Nouvelle paire sélectionnée : {self.pair}")
        else:
            print("Aucune paire disponible n'est ouverte. Veuillez vérifier les heures de marché.")

    def fetch_market_data(self, symbol, period, count=50):
        candles = self.api.get_candles(symbol, period, count, time.time())
        df = pd.DataFrame(candles)
        df['timestamp'] = pd.to_datetime(df['from'], unit='s')
        df.set_index('timestamp', inplace=True)
        return df

    def check_trade_result(self, api, order_id):
        max_attempts = 5
        attempt = 0
        while attempt < max_attempts:
            if self.action != "default":
                result, profit = api.check_win_digital_v2(order_id)
                if result is not None and profit is not None:
                    self.profit = round(profit, 2)
                    self.somme_marge_perdu += self.profit
                    if profit > 0:
                        self.profit_aff = self.profit / self.current_stake
                        self.total_profit += self.profit_aff
                        self.result = "win"
                    else:
                        self.result = "loss"
                    return True
            attempt += 1
            time.sleep(1)
        return False

    def calculate_sma_21(self, df, window):
        df['SMA_21'] = df['close'].rolling(window=window).mean()
        df.fillna(0, inplace=True)
        return df

    def calculate_sma_9(self, df, window):
        df['SMA_9'] = df['close'].rolling(window=window).mean()
        df.fillna(0, inplace=True)
        return df

    def calculate_rsi(self, df, window=5):
        df['RSI'] = ta.momentum.rsi(df['close'], window=window)
        return df

    def run_trading(self):
        action = "PUT"
        try:
            api = IQ_Option(self.email, self.password)
            api.connect()
            api.change_balance("PRACTICE" if self.account_type == "demo" else "REAL")

            balance = api.get_balance()
            if balance is None or balance <= 0:
                raise ValueError("Solde du compte insuffisant ou non disponible.")
            else:
                self.solde_initial = balance

            self.status = "Robot en cours d'exécution"
            martingale_multiplier = self.martingale
            last_order_id = None

            while True:

                if not self.is_pair_open(self.pair):
                    print(f"La paire {self.pair} est fermée. Sélection d'une nouvelle paire...")
                    self.select_new_pair()

                current_time = datetime.datetime.now()
                next_minute = (current_time + datetime.timedelta(minutes=1)).replace(second=0, microsecond=0)
                sleep_duration = (next_minute - current_time).total_seconds()
                time.sleep(sleep_duration)

                try:
                    candles = api.get_candles(self.pair, 60, 50, time.time())
                    if not candles:
                        print("Erreur lors de la récupération des données de marché. Réessayez...")
                        continue
                except Exception as e:
                    print(Fore.RED + f"Erreur lors de la récupération des données de marché : {e}" + Style.RESET_ALL)
                    continue
                df = self.fetch_market_data(self.pair, 60, 50)

                # Vérification et renommage des colonnes nécessaires
                required_columns = {'max': 'high', 'min': 'low', 'close': 'close'}
                df.rename(columns=required_columns, inplace=True)

                if 'high' not in df.columns or 'low' not in df.columns or 'close' not in df.columns:
                    raise ValueError("Colonnes nécessaires (high, low, close) manquantes dans le DataFrame.")

                last_candle = candles[0]

                if self.id != 0:
                    if self.check_trade_result(api, self.id):
                        if self.result == "win":
                            self.initial_testing_action = "default"
                            if self.action == "call":
                                action_color = Fore.GREEN + f"{self.action_}" + Style.RESET_ALL
                            else:
                                action_color = Fore.RED + f"{self.action_}" + Style.RESET_ALL
                            result_color = Fore.GREEN + "WIN" + Style.RESET_ALL
                            profit_color = Fore.GREEN + f"{round(self.profit_aff, 2)}" + Style.RESET_ALL
                            pair_color = Fore.YELLOW + f"{self.pair}" + Style.RESET_ALL
                            total_color = Fore.BLUE + f"{self.total_profit:.2f}" + Style.RESET_ALL
                            print(f"|{self.time}    |  {pair_color}   |    {action_color}    |   {self.current_stake:.2f}    | {self.recuperation_martingale}  |   {result_color}    |    {profit_color}    | {total_color} ")
                            self.current_stake = self.stake
                            self.recuperation_martingale = 0
                        elif self.result == "loss":
                            self.current_stake *= martingale_multiplier
                            self.recuperation_martingale += 1
                df = self.calculate_sma_21(df, 21)
                df = self.calculate_sma_9(df, 9)
                df = self.calculate_rsi(df, 14)

                if self.recuperation_martingale <= 2:
                    if df['close'].iloc[-1] < df['open'].iloc[-1]:
                        self.action = "put"
                        self.action_ = "PUT "
                    elif df['close'].iloc[-1] > df['open'].iloc[-1]:
                        self.action = "call"
                        self.action_ = "CALL"
                else:
                    if df['SMA_9'].iloc[-1] < df['SMA_21'].iloc[-1] and df['RSI'].iloc[-1] < 50:
                        self.action = "put"
                        self.action_ = "PUT "
                    elif df['SMA_9'].iloc[-1] > df['SMA_21'].iloc[-1] and df['RSI'].iloc[-1] > 50:
                        self.action = "call"
                        self.action_ = "CALL"

                if self.action != "default":
                    try:
                        result, order_id = api.buy_digital_spot(self.pair, self.current_stake, self.action, 1)
                        self.id = order_id
                    except Exception as e:
                        self.status = f"Exception: {e}"

                trade_time = datetime.datetime.now()
                formatted_trade_time = trade_time.strftime('%Y-%m-%d %H:%M')
                self.time = formatted_trade_time

                # Append trade details to Excel sheet
                if self.action != "default":
                    row_data = [self.time, self.pair, self.action, self.current_stake, self.recuperation_martingale, self.result, self.profit, self.total_profit]
                    next_row = self.ws.max_row + 1
                    for col_num, value in enumerate(row_data, start=1):
                        cell = self.ws.cell(row=next_row, column=col_num)
                        cell.value = value

                    # Save the workbook after each trade (optional)
                    self.wb.save(self.excel_filename)

        except Exception as e:
            self.status = f"Erreur: {e}"
        finally:
            api.api.close()
            self.wb.save(self.excel_filename)
            self.wb.close()

    def start_trading(self):
        threading.Thread(target=self.run_trading).start()

if __name__ == "__main__":
    bot = TradingBot()
    bot.start_trading()
