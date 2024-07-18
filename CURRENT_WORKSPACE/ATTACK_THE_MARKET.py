import datetime
import time
import openpyxl
import configparser
import random
import logging
import smtplib
from email.mime.text import MIMEText
from iqoptionapi.stable_api import IQ_Option

# Configure logging
logging.basicConfig(filename='trading_log.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def load_config():
    # Load configuration from a file
    config = configparser.ConfigParser()
    config.read('config.ini')
    return config

def connect_to_iqoption(email, password):
    # Connect to the IQ Option API
    api = IQ_Option(email, password)
    status, reason = api.connect()
    return api, status, reason

def send_email(subject, body, smtp_server, smtp_port, smtp_user, smtp_password, notification_email):
    # Function to send email notifications
    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = smtp_user
    msg['To'] = notification_email

    try:
        with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
            server.login(smtp_user, smtp_password)
            server.sendmail(smtp_user, notification_email, msg.as_string())
        logging.info(f"Notification sent: {subject}")
    except Exception as e:
        logging.error(f"Failed to send email notification: {e}")

def get_demo_balance(api, demo_balance):
    # Function to calculate demo balance
    return demo_balance - api.get_balance()

def check_trade_result(api, trade_id, max_retries=3, retry_delay=10):
    # Function to check trade result with retry mechanism
    retries = 0
    while retries < max_retries:
        try:
            trade_result = api.check_win_v3(trade_id)
            if trade_result is not None:
                return trade_result
            else:
                logging.warning("Failed to retrieve trade result, retrying...")
                time.sleep(retry_delay)
        except Exception as e:
            logging.error(f"Error checking trade result: {e}, retrying...")
            time.sleep(retry_delay)
        retries += 1
    logging.error("Max retries exceeded, unable to retrieve trade result")
    return None

def main():
    config = load_config()

    # IQ Option credentials and trading parameters
    email = config['IQOption']['email']
    password = config['IQOption']['password']
    global_amount = float(config['Trading']['amount'])
    martingale = float(config['Trading']['martingale'])
    demo_balance = float(config['Trading']['demo_initial_balance'])
    short_period = int(config['Trading']['short_period'])
    long_period = int(config['Trading']['long_period'])
    asset = config['Trading']['asset']
    duration = int(config['Trading']['duration'])

    # Email configuration for notifications
    smtp_server = config['Email']['smtp_server']
    smtp_port = int(config['Email']['smtp_port'])
    smtp_user = config['Email']['smtp_user']
    smtp_password = config['Email']['smtp_password']
    notification_email = config['Email']['notification_email']

    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Trade Monitoring'

    # Headers for the table
    headers = ['Trade ID', 'Timestamp', 'Strategy', 'Status', 'Amount', 'Trade Result', 'Balance']
    # Write headers to the first row
    sheet.append(headers)

    # Generate a random number for the filename
    random_number = random.randint(1000, 9999)
    excel_file = f'trade_monitoring_{random_number}.xlsx'

    api, status, reason = connect_to_iqoption(email, password)

    if not status:
        logging.error(f'Failed to connect: {reason}')
        print('Failed to connect:', reason)
        exit()

    # If 2FA is enabled
    if reason == "2FA":
        code_sms = input("Enter the received code: ")
        status, reason = api.connect_2fa(code_sms)
        if not status:
            logging.error(f'Failed to connect with 2FA: {reason}')
            print('Failed to connect with 2FA:', reason)
            exit()

    # Real-time trading loop
    while True:
        try:
            # Sleep until the start of the next minute
            current_time = datetime.datetime.now()
            next_minute = (current_time + datetime.timedelta(minutes=1)).replace(second=0, microsecond=0)
            sleep_duration = (next_minute - current_time).total_seconds()
            time.sleep(sleep_duration)

            # Fetch the latest candlestick data
            end_time = time.time()  # Current time
            size = max(short_period, long_period)  # Number of candlesticks needed
            candles = api.get_candles(asset, duration, size, end_time)

            if len(candles) < size:
                logging.warning("Not enough data for analysis")
                print("Not enough data for analysis")
                continue

            # Extract closing prices
            close_prices = [candle['close'] for candle in candles]

            # Calculate moving averages
            short_ma = sum(close_prices[-short_period:]) / short_period if len(close_prices) >= short_period else 0
            long_ma = sum(close_prices[-long_period:]) / long_period if len(close_prices) >= long_period else 0

            # Determine trade direction
            if short_ma > long_ma:
                direction = "call"  # Buy
            else:
                direction = "put"  # Sell

            # Execute the trade
            status, trade_id = api.buy_digital_spot(asset,global_amount,direction, duration)

            if status:
                logging.info(f"Trade executed successfully: {direction} on {asset} with Trade ID {trade_id} and Amount {global_amount}")
                print(f"Trade executed successfully: {direction} on {asset}")

                timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # Current timestamp
                strategy = 'Moving Average Crossover Strategy'  # Strategy name
                trade_status = 'Open'  # Trade status when created

                # Write data to Excel
                sheet.append([trade_id, timestamp, strategy, trade_status, global_amount, None, get_demo_balance(api, demo_balance)])
            else:
                logging.error(f"Trade execution failed: {reason}")
                print(f"Trade execution failed: {reason}")

            # Check the trade result with retry mechanism
            trade_result = check_trade_result(api, trade_id)

            if trade_result is not None:
                result_str = 'Win' if trade_result > 0 else 'Loss'
                logging.info(f"Trade result: {result_str} for Trade ID {trade_id}. Profit/Loss: {trade_result}. New Balance: {get_demo_balance(api, demo_balance)}")
                print(f"Trade result: {result_str}")
                print(f"Profit/Loss: {trade_result}")
                print("Balance:", get_demo_balance(api, demo_balance))

                # Update trade result in the Excel sheet
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    if row[0].value == trade_id:
                        row[5].value = 'Win' if trade_result > 0 else 'Loss'
                        row[6].value = get_demo_balance(api, demo_balance)
                        break

                # Apply Martingale strategy
                if trade_result <= 0:
                    # Check if martingale is supported by the balance
                    if 2 * (martingale * global_amount) <= get_demo_balance(api, demo_balance):
                        global_amount *= martingale
                    else:
                        global_amount = float(config['Trading']['amount'])

                    logging.info(f"Martingale applied. New amount for the next trade: {global_amount}")
                else:
                    global_amount = float(config['Trading']['amount'])
                    logging.info(f"Trade was successful. Resetting amount to {global_amount} for the next trade.")

            else:
                logging.error("Failed to retrieve trade result")
                print("Failed to retrieve trade result")

            # Save workbook to Excel file
            workbook.save(filename=excel_file)
            logging.info(f"Workbook saved as {excel_file}")

        except Exception as e:
            logging.error(f"An error occurred: {e}")
            print(f"An error occurred: {e}")
            time.sleep(10)  # Wait before retrying in case of an error

if __name__ == "__main__":
    main()
